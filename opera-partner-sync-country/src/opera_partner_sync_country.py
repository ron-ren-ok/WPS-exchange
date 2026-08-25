"""Sync Opera country metrics from Looker ZIP reports received by Gmail."""
import argparse
import csv
import email
import imaplib
import io
import json
import os
import re
import sys
import zipfile
from datetime import datetime, timedelta
from email.utils import parseaddr
from zoneinfo import ZoneInfo

SHEET_ID = "1vSBU84SFoVlXdaczYYAev8mC0PEfjRQyVSv8s2OAGW4"
SHEET_NAME = "合作方新增血量分国家"
HEADERS = ("日期", "合作方", "国家代码", "运营位", "新增", "血量")
PARTNER = "Opera"
EXPECTED_MAILBOX = "54lingbai@gmail.com"
SENDER = "noreply@lookermail.com"
SUBJECT = "Opera for Computers distribution partner - download"
SURFACES = {"wpstest2/opera.exe": "换量弹窗", "wpstest": "气泡"}
ALIASES = {
    "date": ("date", "day"),
    "campaign": ("campaign",),
    "country": ("countrycode", "countrycodeiso", "countryiso", "country"),
    "new_users": ("newusers", "newuser", "installs", "installcount"),
    "blood_volume": ("revenue", "earnings", "payout", "commission", "bloodvolume"),
}


def normalize(value):
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def parse_day(value):
    text = str(value).strip().split(",", 1)[0]
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    try:
        serial = float(text)
        if 20000 <= serial <= 80000:
            return (datetime(1899, 12, 30) + timedelta(days=serial)).date()
    except ValueError:
        pass
    raise ValueError(f"unsupported date: {value!r}")


def number(value):
    text = str(value).replace("\u00a0", "").replace(",", "").replace("$", "").strip()
    if not re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        raise ValueError(f"invalid numeric value: {value!r}")
    parsed = float(text)
    return int(parsed) if parsed.is_integer() else parsed


def map_columns(row):
    names = {normalize(header): header for header in row if header not in (None, "")}
    result = {field: next((names[name] for name in aliases if name in names), None)
              for field, aliases in ALIASES.items()}
    missing = [field for field, name in result.items() if name is None]
    if missing:
        raise RuntimeError("Opera export missing required columns: " + ", ".join(missing))
    return result


def csv_records(raw, filename):
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            rows = list(csv.DictReader(io.StringIO(raw.decode(encoding))))
            if rows:
                return rows
        except UnicodeDecodeError:
            pass
    raise RuntimeError(f"could not decode CSV: {filename}")


def report_rows(raw_zip):
    try:
        archive = zipfile.ZipFile(io.BytesIO(raw_zip))
    except zipfile.BadZipFile as exc:
        raise RuntimeError("Opera attachment is not a valid ZIP archive") from exc
    rows = []
    with archive:
        entries = [entry for entry in archive.infolist() if not entry.is_dir()]
        if not entries:
            raise RuntimeError("Opera ZIP has no data files")
        for entry in entries:
            raw = archive.read(entry)
            if entry.filename.lower().endswith(".csv"):
                rows.extend(csv_records(raw, entry.filename))
            elif entry.filename.lower().endswith(".xlsx"):
                from openpyxl import load_workbook
                workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                try:
                    values = workbook.active.iter_rows(values_only=True)
                    headers = next(values, None)
                    if not headers:
                        raise RuntimeError(f"XLSX has no headers: {entry.filename}")
                    rows.extend(dict(zip(headers, row)) for row in values
                                if any(value not in (None, "") for value in row))
                finally:
                    workbook.close()
            else:
                raise RuntimeError(f"unsupported file in Opera ZIP: {entry.filename}")
    return rows


def parse_report(raw_zip, start, end):
    result = {}
    for row in report_rows(raw_zip):
        columns = map_columns(row)
        operation = SURFACES.get(str(row[columns["campaign"]] or "").strip())
        if not operation:
            continue
        date = parse_day(row[columns["date"]])
        if not start <= date <= end:
            continue
        country = str(row[columns["country"]] or "").strip().upper()
        if not re.fullmatch(r"[A-Z]{2}", country):
            raise RuntimeError(f"Opera returned invalid country code: {country!r}")
        record = result.setdefault((date, country, operation), {"new_users": 0, "blood_volume": 0})
        record["new_users"] += number(row[columns["new_users"]])
        record["blood_volume"] += number(row[columns["blood_volume"]])
    if not result:
        raise RuntimeError("Opera report has no mapped rows in the requested date range")
    return result


def gmail_client(username, password):
    if username.strip().lower() != EXPECTED_MAILBOX:
        raise RuntimeError(f"GMAIL_IMAP_USERNAME must be {EXPECTED_MAILBOX}")
    try:
        client = imaplib.IMAP4_SSL("imap.gmail.com", 993)
        client.login(username.strip(), password.replace(" ", "").strip())
        return client
    except imaplib.IMAP4.error as exc:
        raise RuntimeError("Gmail IMAP login failed") from exc


def zip_attachments(client):
    status, mailboxes = client.list()
    if status != "OK":
        raise RuntimeError("Gmail mailbox listing failed")
    all_mail = next((item.decode("utf-8", "replace").rsplit('"', 2)[-2]
                     for item in mailboxes if b"\\All" in item), None)
    if client.select(all_mail or "INBOX", readonly=True)[0] != "OK":
        raise RuntimeError("Gmail mailbox select failed")
    status, data = client.uid("search", None, "SUBJECT", f'"{SUBJECT}"')
    if status != "OK":
        raise RuntimeError("Gmail subject search failed")
    for uid in reversed(data[0].split()):
        status, payload = client.uid("fetch", uid, "(RFC822)")
        if status != "OK" or not payload or not isinstance(payload[0], tuple):
            continue
        message = email.message_from_bytes(payload[0][1])
        if parseaddr(message.get("From", ""))[1].lower() != SENDER:
            continue
        if message.get("Subject", "").strip() != SUBJECT:
            continue
        for part in message.walk():
            raw, filename = part.get_payload(decode=True), part.get_filename() or ""
            if raw and filename.lower().endswith(".zip"):
                yield raw


def sheet_service(raw):
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    credentials = Credentials.from_service_account_info(
        json.loads(raw), scopes=["https://www.googleapis.com/auth/spreadsheets"])
    return build("sheets", "v4", credentials=credentials, cache_discovery=False)


def value_at(row, index):
    return row[index] if index < len(row) else ""


def column_name(index):
    output = ""
    while True:
        index, remainder = divmod(index, 26)
        output = chr(65 + remainder) + output
        if index == 0:
            return output
        index -= 1


def existing_rows(service):
    values = service.spreadsheets().values().get(
        spreadsheetId=SHEET_ID, range=f"'{SHEET_NAME}'!A1:F10000",
        valueRenderOption="UNFORMATTED_VALUE", dateTimeRenderOption="SERIAL_NUMBER",
    ).execute().get("values", [])
    if not values or len(values[0]) != len(set(values[0])) or any(header not in values[0] for header in HEADERS):
        raise RuntimeError("country-detail target headers are missing or duplicated")
    headers = values[0]
    positions = {header: headers.index(header) for header in HEADERS}
    result = {}
    for row_number, row in enumerate(values[1:], start=2):
        if not row or not value_at(row, positions["日期"]):
            continue
        key = (parse_day(value_at(row, positions["日期"])),
               str(value_at(row, positions["合作方"])).strip(),
               str(value_at(row, positions["国家代码"])).strip().upper(),
               str(value_at(row, positions["运营位"])).strip())
        if key in result:
            raise RuntimeError(f"duplicate country-detail record: {key}")
        result[key] = (row_number, row)
    return headers, result


def same_value(current, wanted):
    try:
        return abs(float(current) - float(wanted)) < 1e-9
    except (ValueError, TypeError):
        return str(current).replace(",", "") == str(wanted)


def plan_writes(headers, target, source, overwrite):
    positions = {header: headers.index(header) for header in HEADERS}
    updates, appends, conflicts = [], [], []
    for (date, country, operation), metrics in sorted(source.items()):
        found = target.get((date, PARTNER, country, operation))
        if found is None:
            appends.append({"日期": date, "合作方": PARTNER, "国家代码": country, "运营位": operation,
                            "新增": metrics["new_users"], "血量": metrics["blood_volume"]})
            continue
        row_number, row = found
        for header, metric in (("新增", "new_users"), ("血量", "blood_volume")):
            current, wanted = value_at(row, positions[header]), metrics[metric]
            if current in ("", None) or not same_value(current, wanted):
                if current not in ("", None) and not overwrite:
                    conflicts.append(f"{date} Opera/{country}/{operation}/{header}")
                else:
                    updates.append({"range": f"'{SHEET_NAME}'!{column_name(positions[header])}{row_number}",
                                    "values": [[wanted]]})
    if conflicts:
        raise RuntimeError("refusing to overwrite country-detail conflicts: " + "; ".join(conflicts))
    return updates, appends


def append_rows(service, headers, records):
    if not records:
        return
    positions = {header: headers.index(header) for header in HEADERS}
    values = []
    for record in records:
        row = [""] * len(headers)
        for header in HEADERS:
            row[positions[header]] = record[header].isoformat() if header == "日期" else record[header]
        values.append(row)
    service.spreadsheets().values().append(
        spreadsheetId=SHEET_ID, range=f"'{SHEET_NAME}'!A1:{column_name(len(headers) - 1)}",
        valueInputOption="USER_ENTERED", insertDataOption="INSERT_ROWS",
        body={"majorDimension": "ROWS", "values": values},
    ).execute()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--start-date")
    parser.add_argument("--end-date")
    parser.add_argument("--allow-overwrite", action="store_true")
    args = parser.parse_args()
    username = os.environ.get("GMAIL_IMAP_USERNAME", "")
    password = os.environ.get("GMAIL_APP_PASSWORD", "")
    service_json = os.environ.get("GOOGLE_SHEET_SERVICE_ACCOUNT_JSON", "")
    if not username or not password or not service_json:
        raise RuntimeError("missing required GitHub Actions secret")
    end = parse_day(args.end_date) if args.end_date else datetime.now(ZoneInfo("Asia/Shanghai")).date() - timedelta(days=1)
    start = parse_day(args.start_date) if args.start_date else end
    if start > end:
        raise RuntimeError("start date is after end date")
    gmail = gmail_client(username, password)
    try:
        source = {}
        for raw_zip in zip_attachments(gmail):
            for key, metrics in parse_report(raw_zip, start, end).items():
                source.setdefault(key, metrics)
    finally:
        gmail.logout()
    if not source:
        raise RuntimeError("no verified Opera country-report ZIP attachment was found")
    service = sheet_service(service_json)
    headers, target = existing_rows(service)
    updates, appends = plan_writes(headers, target, source, args.allow_overwrite)
    if updates:
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=SHEET_ID, body={"valueInputOption": "USER_ENTERED", "data": updates}).execute()
    append_rows(service, headers, appends)
    print(json.dumps({"records": len(source), "updated_cells": len(updates),
                      "appended_rows": len(appends)}, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except (RuntimeError, ValueError, OSError, json.JSONDecodeError, zipfile.BadZipFile) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(2)
